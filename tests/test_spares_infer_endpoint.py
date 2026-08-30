# tests/test_spares_infer_endpoint.py — POST /spares/infer, the endpoint that wires
# the column-inference engine (already unit-tested in test_spares_column_inference.py)
# to a real file upload: CSV parsing, xlsx parsing via openpyxl (bold-based category
# row detection, duplicate/blank header handling, cell normalization), and the
# empty-file/no-data-rows/generic-failure error paths. This whole function (spares.py
# lines 313-439) had zero coverage before this file despite being real, error-prone
# spreadsheet-import logic used for bulk spare uploads.

import io

import pytest
from fastapi import HTTPException
from openpyxl import Workbook

from app.routers.spares import infer_spare_columns


class _FakeUploadFile:
    def __init__(self, filename: str, content: bytes):
        self.filename = filename
        self._content = content

    async def read(self) -> bytes:
        return self._content


def _xlsx_bytes(rows: list, bold_rows: set = frozenset()) -> bytes:
    """Build a real .xlsx file. `rows` is a list of row-tuples (row 0 = header).
    `bold_rows` is a set of row indices (0-based) whose cells should be bold."""
    wb = Workbook()
    ws = wb.active
    for r_idx, row in enumerate(rows):
        ws.append(row)
        if r_idx in bold_rows:
            for cell in ws[r_idx + 1]:
                if cell.value is not None:
                    cell.font = cell.font.copy(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── CSV path ─────────────────────────────────────────────────────────────────────

async def test_infer_csv_happy_path_returns_inferred_roles_and_raw_rows():
    content = (
        "Stock Code,Description,Unit Price\n"
        "AB-1234,Hydraulic hose fitting brass connector,12.50\n"
        "CD-5678,Ball bearing 6205 2RS sealed unit,8.99\n"
    ).encode()
    file = _FakeUploadFile("parts.csv", content)
    result = await infer_spare_columns(file=file, current_user={})
    assert result["inferred"]["stock_code"] == "Stock Code"
    assert result["inferred"]["description"] == "Description"
    assert result["inferred"]["unit_price"] == "Unit Price"
    assert result["total_rows"] == 2
    assert result["has_categories"] is False
    assert len(result["raw_rows"]) == 2
    assert result["raw_rows"][0]["Stock Code"] == "AB-1234"


async def test_infer_csv_header_only_is_no_data_rows_400():
    content = "Stock Code,Description,Unit Price\n".encode()
    file = _FakeUploadFile("parts.csv", content)
    with pytest.raises(HTTPException) as exc:
        await infer_spare_columns(file=file, current_user={})
    assert exc.value.status_code == 400


async def test_infer_tsv_extension_uses_tab_separator():
    content = "Stock Code\tDescription\tUnit Price\nAB-1\tWidget\t5.00\n".encode()
    file = _FakeUploadFile("parts.tsv", content)
    result = await infer_spare_columns(file=file, current_user={})
    assert result["total_rows"] == 1
    assert result["raw_rows"][0]["Stock Code"] == "AB-1"


# ─── xlsx path ────────────────────────────────────────────────────────────────────

async def test_infer_xlsx_happy_path():
    content = _xlsx_bytes([
        ("Stock Code", "Description", "Unit Price"),
        ("AB-1234", "Hydraulic hose fitting brass connector", 12.50),
        ("CD-5678", "Ball bearing 6205 2RS sealed unit", 8.99),
    ])
    file = _FakeUploadFile("parts.xlsx", content)
    result = await infer_spare_columns(file=file, current_user={})
    assert result["total_rows"] == 2
    assert result["inferred"]["unit_price"] == "Unit Price"


async def test_infer_xlsx_bold_category_row_is_extracted_not_a_data_row():
    content = _xlsx_bytes([
        ("Stock Code", "Description", "Unit Price"),
        ("Bearings", None, None),
        ("AB-1234", "Ball bearing sealed unit", 8.99),
    ], bold_rows={1})
    file = _FakeUploadFile("parts.xlsx", content)
    result = await infer_spare_columns(file=file, current_user={})
    assert result["total_rows"] == 1
    assert result["has_categories"] is True
    assert result["raw_rows"][0]["_category"] == "Bearings"


async def test_infer_xlsx_category_prefix_row_extracted_even_without_bold():
    content = _xlsx_bytes([
        ("Stock Code", "Description", "Unit Price"),
        ("Category: Hydraulics", None, None),
        ("AB-1234", "Hydraulic hose fitting brass", 12.50),
    ])
    file = _FakeUploadFile("parts.xlsx", content)
    result = await infer_spare_columns(file=file, current_user={})
    assert result["total_rows"] == 1
    assert result["raw_rows"][0]["_category"] == "Hydraulics"


async def test_infer_xlsx_duplicate_headers_get_suffixed():
    content = _xlsx_bytes([
        ("Code", "Code", "Unit Price"),
        ("AB-1234", "AB-1234-alt", 12.50),
    ])
    file = _FakeUploadFile("parts.xlsx", content)
    result = await infer_spare_columns(file=file, current_user={})
    assert "Code" in result["all_columns"]
    assert "Code_1" in result["all_columns"]


async def test_infer_xlsx_blank_header_gets_placeholder_name():
    content = _xlsx_bytes([
        ("Stock Code", None, "Unit Price"),
        ("AB-1234", "extra", 12.50),
    ])
    file = _FakeUploadFile("parts.xlsx", content)
    result = await infer_spare_columns(file=file, current_user={})
    assert "Column_2" in result["all_columns"]


async def test_infer_xlsx_blank_rows_are_skipped():
    content = _xlsx_bytes([
        ("Stock Code", "Description", "Unit Price"),
        (None, None, None),
        ("AB-1234", "Widget", 12.50),
    ])
    file = _FakeUploadFile("parts.xlsx", content)
    result = await infer_spare_columns(file=file, current_user={})
    assert result["total_rows"] == 1


async def test_infer_xlsx_no_data_rows_is_400():
    content = _xlsx_bytes([("Stock Code", "Description", "Unit Price")])
    file = _FakeUploadFile("parts.xlsx", content)
    with pytest.raises(HTTPException) as exc:
        await infer_spare_columns(file=file, current_user={})
    assert exc.value.status_code == 400


async def test_infer_drops_entirely_null_columns_before_inference():
    content = _xlsx_bytes([
        ("Stock Code", "Empty Col", "Unit Price"),
        ("AB-1234", None, 12.50),
        ("CD-5678", None, 8.99),
    ])
    file = _FakeUploadFile("parts.xlsx", content)
    result = await infer_spare_columns(file=file, current_user={})
    assert "Empty Col" not in result["all_columns"]


# ─── error handling ───────────────────────────────────────────────────────────────

async def test_infer_xlsx_truly_empty_workbook_is_400():
    # No rows at all (not even a header row) -- distinct from "header row but no data
    # rows", which is the separate has-data-rows-but-empty-df 400 case.
    wb = Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    file = _FakeUploadFile("parts.xlsx", buf.getvalue())
    with pytest.raises(HTTPException) as exc:
        await infer_spare_columns(file=file, current_user={})
    assert exc.value.status_code == 400


async def test_infer_rejects_unsupported_extension():
    file = _FakeUploadFile("parts.exe", b"whatever")
    with pytest.raises(HTTPException) as exc:
        await infer_spare_columns(file=file, current_user={})
    assert exc.value.status_code == 400


async def test_infer_generic_failure_is_500_not_swallowed():
    class _BrokenUploadFile:
        filename = "parts.csv"
        async def read(self):
            raise OSError("stream reset by peer")

    with pytest.raises(HTTPException) as exc:
        await infer_spare_columns(file=_BrokenUploadFile(), current_user={})
    assert exc.value.status_code == 500
